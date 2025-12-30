from flask import Flask, render_template, request, jsonify, send_file
import sqlite3, os, re
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
DB = "database.db"
EXCEL = "cadastros.xlsx"

# ================= CPF VALIDAÇÃO MATEMÁTICA =================
def cpf_valido(cpf):
    if not cpf.isdigit() or len(cpf) != 11:
        return False
    if cpf == cpf[0] * 11:
        return False
    for i in range(9, 11):
        soma = sum(int(cpf[num]) * ((i + 1) - num) for num in range(i))
        dig = (soma * 10 % 11) % 10
        if dig != int(cpf[i]):
            return False
    return True

# ================= BANCO =================
def conectar():
    return sqlite3.connect(DB)

with conectar() as con:
    con.execute("""
    CREATE TABLE IF NOT EXISTS beneficiarios (
        cpf TEXT PRIMARY KEY,
        nome TEXT,
        profissao TEXT,
        atividade TEXT,
        renda REAL,
        estado_civil TEXT,
        beneficio TEXT,
        endereco TEXT,
        telefone TEXT,
        pcd INTEGER,
        idosos INTEGER,
        criancas INTEGER,
        moradores INTEGER,
        conjuge_nome TEXT,
        conjuge_cpf TEXT,
        conjuge_profissao TEXT,
        conjuge_atividade TEXT,
        conjuge_renda REAL,
        data TEXT
    )
    """)

# ================= PDF =================
def gerar_pdf(d):
    nome_pdf = f"{d['nome']}.pdf"
    c = canvas.Canvas(nome_pdf, pagesize=A4)
    w, h = A4
    y = h - 50

    def titulo(t):
        nonlocal y
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(w/2, y, t)
        y -= 30

    def secao(t):
        nonlocal y
        c.setFont("Helvetica-Bold", 11)
        c.drawString(40, y, t)
        y -= 18

    def linha(l, v):
        nonlocal y
        c.setFont("Helvetica", 10)
        c.drawString(40, y, f"{l} {v or ''}")
        y -= 14

    titulo("FICHA DE CADASTRO – MCMV RURAL")

    secao("Dados do Beneficiário")
    linha("Nome:", d["nome"])
    linha("CPF:", d["cpf"])
    linha("Profissão:", d["profissao"])
    linha("Atividade:", d["atividade"])
    linha("Renda:", d["renda"])
    linha("Estado Civil:", d["estado_civil"])
    linha("Benefício:", d["beneficio"])

    if d["estado_civil"] in ["Casado", "União Estável"]:
        secao("Dados do Cônjuge")
        linha("Nome:", d["conjuge_nome"])
        linha("CPF:", d["conjuge_cpf"])
        linha("Profissão:", d["conjuge_profissao"])
        linha("Atividade:", d["conjuge_atividade"])
        linha("Renda:", d["conjuge_renda"])

    secao("Contato")
    linha("Endereço:", d["endereco"])
    linha("Telefone:", d["telefone"])

    secao("Informações Adicionais")
    linha("PCD:", d["pcd"])
    linha("Idosos:", d["idosos"])
    linha("Crianças:", d["criancas"])
    linha("Moradores:", d["moradores"])

    y -= 30
    c.drawRightString(w-40, y, f"Mojuí dos Campos - Pará, {datetime.now().strftime('%d/%m/%Y')}")

    c.save()
    return nome_pdf

# ================= EXCEL =================
def salvar_excel(d):
    if not os.path.exists(EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.append(d.keys())
    else:
        wb = load_workbook(EXCEL)
        ws = wb.active
    ws.append(d.values())
    wb.save(EXCEL)

# ================= ROTAS =================
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/salvar", methods=["POST"])
def salvar():
    d = request.form.to_dict()

    # Validações
    if not re.fullmatch(r"[A-Za-zÀ-ÿ ]+", d.get("nome","")):
        return jsonify({"erro": "Nome inválido"}), 400
    if not cpf_valido(d.get("cpf","")):
        return jsonify({"erro": "CPF inválido"}), 400
    if not d.get("telefone","").isdigit():
        return jsonify({"erro": "Telefone inválido"}), 400
    if not d.get("renda","").replace(".","").isdigit():
        return jsonify({"erro": "Renda inválida"}), 400

    with conectar() as con:
        con.execute("""
        INSERT OR REPLACE INTO beneficiarios VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            d["cpf"], d["nome"], d["profissao"], d["atividade"], float(d["renda"]),
            d["estado_civil"], d["beneficio"], d["endereco"], d["telefone"],
            int(d["pcd"] or 0), int(d["idosos"] or 0), int(d["criancas"] or 0), int(d["moradores"] or 0),
            d.get("conjuge_nome"), d.get("conjuge_cpf"), d.get("conjuge_profissao"),
            d.get("conjuge_atividade"), float(d.get("conjuge_renda") or 0),
            datetime.now().strftime("%d/%m/%Y")
        ))

    salvar_excel(d)
    pdf = gerar_pdf(d)
    return send_file(pdf, as_attachment=True)

@app.route("/consultar/<cpf>")
def consultar(cpf):
    with conectar() as con:
        cur = con.execute("SELECT * FROM beneficiarios WHERE cpf=?", (cpf,))
        r = cur.fetchone()
        if not r:
            return jsonify({"erro": "CPF não encontrado"})
        cols = [c[0] for c in cur.description]
        return jsonify(dict(zip(cols, r)))

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
